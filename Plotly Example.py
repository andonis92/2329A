# -*- coding: utf-8 -*-
"""
Created on Tue Feb 15 14:14:35 2022

@author: M310648
"""

import plotly
from plotly.graph_objs import Scatter, Layout
import pandas as pd
from openpyxl import Workbook, load_workbook
import datetime

# Loading Historical Data
WB1 = load_workbook("2329A Graph.xlsx", data_only = True)
WB1_WS1 = WB1["Sheet1"]

date = []
time_one = []
date_time = []
temperature = []

forcast_index = 0

for i in range(1, WB1_WS1.max_row+1):
    date.append(str(WB1_WS1.cell(column = 1, row = i).value))
    time_one.append(str(WB1_WS1.cell(column = 2, row = i).value))
    temperature.append(WB1_WS1.cell(column = 3, row = i).value)

for items in date:
    date_time.append(items + time_one[forcast_index])
    forcast_index += 1


plotly.offline.plot({
    "data": [Scatter(x=date_time, y=temperature)],
    "layout": Layout(title="2329A")
})

